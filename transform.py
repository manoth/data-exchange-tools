"""
Transform Module - แปลงข้อมูล Excel
อ่านไฟล์ Excel, แปลงข้อมูลโดยจับคู่กับฐานข้อมูล, ส่งออกเป็น Excel ใหม่
"""

import os
import json
import urllib.error
import urllib.request
import socket
from datetime import date, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from database import get_connection
from config import APP_DIR, load_agent_config, save_agent_api_key, clear_agent_api_key

# กำหนด path สำหรับไฟล์อัพโหลด
UPLOADS_DIR = os.path.join(APP_DIR, "uploads")
CENTRAL_API_URL = os.environ.get("CENTRAL_API_URL", "https://apicpho.moph.go.th").rstrip("/")
CENTRAL_API_ENROLLMENT_TOKEN = os.environ.get("CENTRAL_API_ENROLLMENT_TOKEN", "data-exchange-agent-enroll-dev-token")
APP_VERSION = os.environ.get("APP_VERSION", "0.1.3")
APP_PORT = int(os.environ.get("PORT", "8899"))


def _get_agent_uid() -> str:
    return f"data-exchange-tools-{socket.gethostname()}-{APP_PORT}"


def _agent_api_headers() -> dict:
    agent_config = load_agent_config()
    headers = {
        "Content-Type": "application/json",
        "X-Agent-Uid": _get_agent_uid(),
    }
    if agent_config.get("api_key"):
        headers["X-Agent-Key"] = agent_config["api_key"]
    return headers


def _read_hospital_info_for_registration() -> dict:
    """อ่านข้อมูลหน่วยบริการจาก opdconfig เพื่อใช้ลงทะเบียน Agent กับ API Center"""
    connection = None
    try:
        connection = get_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT hospitalcode, hospitalname FROM opdconfig LIMIT 1")
            row = cursor.fetchone() or {}
        return {
            "facility_code": str(row.get("hospitalcode") or "").strip(),
            "facility_name": str(row.get("hospitalname") or "").strip(),
            "db_status": "ok",
        }
    except Exception:
        return {
            "facility_code": "",
            "facility_name": "",
            "db_status": "failed",
        }
    finally:
        try:
            if connection:
                connection.close()
        except Exception:
            pass


def _ensure_agent_api_key_for_lookup(force: bool = False) -> dict:
    """ให้ transform ขอ API key เองได้ กรณี heartbeat ยังไม่ได้ key หรือ key เก่าใช้ไม่ได้"""
    agent_config = load_agent_config()
    saved_api_url = (agent_config.get("api_center_url") or "").rstrip("/")
    if (
        not force
        and agent_config.get("api_key")
        and (not saved_api_url or saved_api_url == CENTRAL_API_URL)
    ):
        return {"configured": True, "message": ""}

    if force or (agent_config.get("api_key") and saved_api_url and saved_api_url != CENTRAL_API_URL):
        clear_agent_api_key()

    hospital = _read_hospital_info_for_registration()
    payload = {
        "agentUid": _get_agent_uid(),
        "facilityCode": hospital["facility_code"],
        "facilityName": hospital["facility_name"],
        "machineName": socket.gethostname(),
        "appVersion": APP_VERSION,
        "frontendVersion": APP_VERSION,
        "dbStatus": hospital["db_status"],
    }
    request = urllib.request.Request(
        f"{CENTRAL_API_URL}/api/agents/register",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "X-Agent-Enrollment-Token": CENTRAL_API_ENROLLMENT_TOKEN,
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=10) as response:
        result = json.loads(response.read().decode("utf-8"))

    data = result.get("data") or {}
    api_key = data.get("apiKey")
    api_key_prefix = data.get("apiKeyPrefix") or ""
    if not api_key:
        raise RuntimeError("API Center ไม่ได้ส่ง Agent API key กลับมา")

    save_result = save_agent_api_key(api_key, api_key_prefix, CENTRAL_API_URL)
    if not save_result.get("success"):
        raise RuntimeError(save_result.get("message") or "บันทึก Agent API key ไม่สำเร็จ")
    return {"configured": True, "message": ""}


def _cell_to_text(value) -> str:
    """แปลงค่า cell เป็น text โดยรักษารหัสที่มีศูนย์นำหน้าเท่าที่ openpyxl ส่งมาได้"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_pid(value) -> str:
    """Normalize pid สำหรับเทียบกับ person.person_id โดยไม่บังคับตัดศูนย์นำหน้า"""
    text = _cell_to_text(value)
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_hoscode(value) -> str:
    """Normalize hospital code จาก opdconfig/excel ให้เทียบกันตรงที่สุด"""
    text = _cell_to_text(value).strip().strip('"').strip("'")
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_column_name(value, index: int) -> str:
    text = _cell_to_text(value) if value is not None else ""
    return (text or f"COLUMN_{index + 1}").strip().upper()


def _make_unique_columns(raw_columns: list) -> list:
    columns = []
    seen = {}
    for index, cell in enumerate(raw_columns):
        base_name = _normalize_column_name(cell, index)
        count = seen.get(base_name, 0) + 1
        seen[base_name] = count
        columns.append(base_name if count == 1 else f"{base_name}_{count}")
    return columns


def _extract_facilities(data: list) -> list:
    facilities = {}
    for row in data:
        hoscode = _cell_to_text(row.get("HOSCODE", ""))
        if not hoscode:
            continue
        hosname = _cell_to_text(row.get("HOSNAME", ""))
        item = facilities.setdefault(hoscode, {"hoscode": hoscode, "hosname": hosname, "rows": 0})
        item["rows"] += 1
        if hosname and not item.get("hosname"):
            item["hosname"] = hosname
    return sorted(facilities.values(), key=lambda item: (item["hoscode"], item["hosname"]))


def _chunks(values: list, size: int = 800):
    for index in range(0, len(values), size):
        yield values[index:index + size]


def _normalize_discharge(value) -> str:
    text = _cell_to_text(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_central_person_key(value) -> str:
    text = _cell_to_text(value).strip()
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    digits = "".join(char for char in text if char.isdigit())
    if not digits:
        return text
    if len(digits) == 12:
        return digits.zfill(13)
    return digits


def _normalize_cid_prefix(value) -> str:
    digits = "".join(char for char in _cell_to_text(value) if char.isdigit())
    return digits[:9] if len(digits) >= 9 else ""


def _normalize_sex(value) -> str:
    text = _cell_to_text(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_birth(value) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = _cell_to_text(value).strip()
    if not text:
        return ""
    text = text.split("T", 1)[0].split(" ", 1)[0].replace("/", "-")
    digits = "".join(char for char in text if char.isdigit())
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
    return text


def _normalize_name_prefix(value) -> str:
    """ลบเครื่องหมายปกปิดแล้ว normalize สำหรับเทียบแบบ SQL LIKE 'prefix%'."""
    return _cell_to_text(value).replace("*", "").strip().casefold()


def _pid_lookup_keys(value) -> list:
    pid = _normalize_pid(value)
    if not pid:
        return []
    keys = [pid]
    if pid.isdigit():
        keys.append(str(int(pid)))
    return list(dict.fromkeys(keys))


def _is_complete_person_cid(value) -> bool:
    return len(_normalize_central_person_key(value)) == 13


def _lookup_central_death_pids(pids: list) -> dict:
    """ถาม API กลางว่ามี PID/CID ใดอยู่ในฐานข้อมูลคนตายกลางแล้วบ้าง"""
    clean_pids = sorted({_normalize_central_person_key(pid) for pid in pids if _normalize_central_person_key(pid)})
    if not clean_pids or not CENTRAL_API_URL:
        return {
            "matched": set(),
            "available": bool(CENTRAL_API_URL),
            "message": "" if CENTRAL_API_URL else "ยังไม่ได้กำหนดเส้นทาง API Center",
        }

    try:
        _ensure_agent_api_key_for_lookup()
    except Exception as exc:
        return {
            "matched": set(),
            "available": False,
            "message": f"ไม่สามารถลงทะเบียน Agent API Key เพื่อเทียบข้อมูลการตายได้ ({exc})",
        }

    matched = set()
    failed_batches = 0
    total_batches = 0
    last_error = ""
    for batch in _chunks(clean_pids, 1000):
        total_batches += 1
        try:
            result = _request_central_death_lookup(batch)
        except urllib.error.HTTPError as exc:
            if exc.code in (401, 403):
                try:
                    _ensure_agent_api_key_for_lookup(force=True)
                    result = _request_central_death_lookup(batch)
                except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, RuntimeError) as retry_exc:
                    failed_batches += 1
                    last_error = str(retry_exc)
                    continue
            else:
                failed_batches += 1
                last_error = str(exc)
                continue
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            failed_batches += 1
            last_error = str(exc)
            continue

        if result.get("ok"):
            data = result.get("data") or {}
            for pid in data.get("matchedPids") or []:
                matched.add(_normalize_central_person_key(pid))

    if total_batches > 0 and failed_batches == total_batches:
        return {
            "matched": set(),
            "available": False,
            "message": f"ไม่สามารถเชื่อมต่อ API Center เพื่อเทียบข้อมูลการตายกับส่วนกลางได้ ({last_error or 'connection failed'})",
        }

    message = ""
    if failed_batches:
        message = "เทียบข้อมูลการตายกับส่วนกลางได้บางส่วน เนื่องจากบางชุดข้อมูลติดต่อ API Center ไม่สำเร็จ"
    return {
        "matched": matched,
        "available": True,
        "message": message,
    }


def _request_central_death_lookup(batch: list) -> dict:
    payload = json.dumps({"pids": batch}).encode("utf-8")
    request = urllib.request.Request(
        f"{CENTRAL_API_URL}/api/agents/death-persons/lookup",
        data=payload,
        headers=_agent_api_headers(),
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=15) as response:
        return json.loads(response.read().decode("utf-8"))


def lookup_central_death_pids(pids: list) -> dict:
    """Public entry point for reports that use the secured central lookup."""
    return _lookup_central_death_pids(pids)


def process_upload(file_path: str) -> dict:
    """
    อ่านไฟล์ Excel และส่งคืนข้อมูล columns, data, preview

    Args:
        file_path: path ไปยังไฟล์ Excel

    Returns:
        dict: {columns, data, preview}
    """
    try:
        # อ่านไฟล์ Excel (read_only=True สำหรับไฟล์ขนาดใหญ่)
        wb = load_workbook(file_path, read_only=True, data_only=True)

        # ใช้ sheet ชื่อ 'Data' หรือ sheet แรก
        if "Data" in wb.sheetnames:
            ws = wb["Data"]
        else:
            ws = wb.active

        # อ่าน header จากแถวแรก
        columns = []
        rows_data = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                # แถวแรกเป็น header
                columns = _make_unique_columns(list(row))
            else:
                # แถวข้อมูล
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(columns):
                        col_name = columns[col_idx]
                        row_dict[col_name] = _cell_to_text(cell)
                # ข้ามแถวว่าง
                if any(v != "" and v is not None for v in row_dict.values()):
                    rows_data.append(row_dict)

        wb.close()

        # สร้าง preview (5 แถวแรก)
        preview = rows_data[:5]

        return {
            "columns": columns,
            "data": rows_data,
            "preview": preview,
            "facilities": _extract_facilities(rows_data)
        }

    except Exception as e:
        raise Exception(f"เกิดข้อผิดพลาดในการอ่านไฟล์ Excel: {e}")


def transform_data(file_id: str, data: list, columns: list, selected_hoscodes: list = None) -> dict:
    """จับคู่ PID ก่อน แล้ว fallback ด้วย CID/เพศ/วันเกิด/ชื่อ/นามสกุล."""
    try:
        column_by_name = {str(col).upper(): col for col in columns}
        father_column = column_by_name.get("FATHER")
        mother_column = column_by_name.get("MOTHER")
        discharge_column = column_by_name.get("DISCHARGE")
        cid_column = column_by_name.get("CID")
        pid_column = column_by_name.get("PID")
        hoscode_column = column_by_name.get("HOSCODE")
        sex_column = column_by_name.get("SEX")
        birth_column = column_by_name.get("BIRTH")
        name_column = column_by_name.get("NAME")
        lname_column = column_by_name.get("LNAME")
        person_columns = ["PERSON_CID", "FULL_NAME", "เงื่อนไขที่ใช้", "เทียบตาย"]

        selected_hoscodes = [str(code).strip() for code in (selected_hoscodes or []) if str(code).strip()]
        if selected_hoscodes:
            selected_set = {_normalize_hoscode(code) for code in selected_hoscodes}
            data = [
                row for row in data
                if _normalize_hoscode(row.get(hoscode_column or "", "")) in selected_set
            ]

        def empty_person_fields() -> dict:
            return {
                "PERSON_CID": "",
                "FULL_NAME": "",
                "เงื่อนไขที่ใช้": "ไม่พบข้อมูล",
                "เทียบตาย": "-",
                "_matched": False,
                "_pid_matched": False,
                "_cid_matched": False,
                "_match_method": "none",
                "_central_death_mismatch": False,
            }

        def apply_person_fields(target: dict, person: dict, method: str) -> None:
            target["PERSON_CID"] = person.get("cid", "")
            target["FULL_NAME"] = person.get("full_name", "")
            if father_column and person.get("father_name"):
                target[father_column] = person.get("father_name", "")
            if mother_column and person.get("mother_name"):
                target[mother_column] = person.get("mother_name", "")
            target["_matched"] = True
            target["_pid_matched"] = method == "pid"
            target["_cid_matched"] = method == "cid"
            target["_match_method"] = method
            target["เงื่อนไขที่ใช้"] = "PID" if method == "pid" else "CID"
            target["_central_death_mismatch"] = False

        pid_values = set()
        cid_prefixes = set()
        for row in data:
            pid_values.update(_pid_lookup_keys(row.get(pid_column or "", "")))
            prefix = _normalize_cid_prefix(row.get(cid_column or "", ""))
            if prefix:
                cid_prefixes.add(prefix)

        results = []
        local_hoscodes = set()
        connection = None
        try:
            connection = get_connection()
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT hospitalcode
                    FROM opdconfig
                    WHERE hospitalcode IS NOT NULL AND hospitalcode <> ''
                """)
                local_hoscodes = {
                    _normalize_hoscode(row.get("hospitalcode"))
                    for row in cursor.fetchall()
                    if _normalize_hoscode(row.get("hospitalcode"))
                }

                for batch in _chunks(sorted(pid_values)):
                    placeholders = ",".join(["%s"] * len(batch))
                    sql = f"""
                        SELECT
                            person_id,
                            cid,
                            pname,
                            fname,
                            lname,
                            father_name,
                            mother_name,
                            sex,
                            birthdate
                        FROM person
                        WHERE CAST(person_id AS CHAR) IN ({placeholders})
                    """
                    cursor.execute(sql, tuple(batch))
                    results.extend(cursor.fetchall())

                for batch in _chunks(sorted(cid_prefixes)):
                    placeholders = ",".join(["%s"] * len(batch))
                    sql = f"""
                        SELECT
                            person_id,
                            cid,
                            pname,
                            fname,
                            lname,
                            father_name,
                            mother_name,
                            sex,
                            birthdate
                        FROM person
                        WHERE LEFT(CAST(cid AS CHAR), 9) IN ({placeholders})
                    """
                    cursor.execute(sql, tuple(batch))
                    results.extend(cursor.fetchall())
        finally:
            if connection:
                connection.close()

        person_map = {}
        cid_match_map = {}
        for row in results:
            person_id = _normalize_pid(row.get("person_id"))
            person = {
                "cid": row.get("cid", ""),
                "full_name": "".join([
                    str(row.get("pname") or ""),
                    str(row.get("fname") or ""),
                    " ",
                    str(row.get("lname") or ""),
                ]).strip(),
                "father_name": row.get("father_name", "") or "",
                "mother_name": row.get("mother_name", "") or "",
                "sex": _normalize_sex(row.get("sex")),
                "birth": _normalize_birth(row.get("birthdate")),
                "fname": _normalize_name_prefix(row.get("fname")),
                "lname": _normalize_name_prefix(row.get("lname")),
            }
            for key in _pid_lookup_keys(person_id):
                person_map[key] = person
            cid_key = (
                _normalize_cid_prefix(row.get("cid")),
                person["sex"],
                person["birth"],
            )
            if all(cid_key):
                cid_match_map.setdefault(cid_key, []).append(person)

        pid_matched_count = 0
        cid_matched_count = 0
        central_death_candidates = {}
        transformed_data = []

        for row in data:
            new_row = dict(row)
            new_row.update(empty_person_fields())
            row_hoscode = _normalize_hoscode(row.get(hoscode_column, ""))
            hoscode_matched = bool(row_hoscode and row_hoscode in local_hoscodes)
            person = None

            if pid_column and hoscode_column and hoscode_matched:
                person = next(
                    (person_map[key] for key in _pid_lookup_keys(row.get(pid_column, "")) if key in person_map),
                    None,
                )
            if person:
                apply_person_fields(new_row, person, "pid")
                pid_matched_count += 1
            elif cid_column and sex_column and birth_column and name_column and lname_column:
                cid_key = (
                    _normalize_cid_prefix(row.get(cid_column, "")),
                    _normalize_sex(row.get(sex_column, "")),
                    _normalize_birth(row.get(birth_column, "")),
                )
                excel_fname_prefix = _normalize_name_prefix(row.get(name_column, ""))
                excel_lname_prefix = _normalize_name_prefix(row.get(lname_column, ""))
                candidates = cid_match_map.get(cid_key, []) if all(cid_key) else []
                person = next(
                    (
                        candidate for candidate in candidates
                        if candidate["fname"].startswith(excel_fname_prefix)
                        and candidate["lname"].startswith(excel_lname_prefix)
                    ),
                    None,
                )
                if person:
                    apply_person_fields(new_row, person, "cid")
                    cid_matched_count += 1

            if (
                discharge_column
                and new_row["_matched"]
                and _normalize_discharge(row.get(discharge_column, "")) != "1"
                and _is_complete_person_cid(new_row.get("PERSON_CID"))
            ):
                death_lookup_pid = _normalize_central_person_key(new_row["PERSON_CID"])
                new_row["เทียบตาย"] = "รอตรวจสอบ"
                central_death_candidates.setdefault(death_lookup_pid, []).append(new_row)

            transformed_data.append(new_row)

        matched_count = pid_matched_count + cid_matched_count
        unmatched_count = len(transformed_data) - matched_count
        central_death_mismatch_count = 0
        central_death_lookup_available = True
        central_death_lookup_message = ""
        if central_death_candidates:
            death_lookup_result = _lookup_central_death_pids(list(central_death_candidates.keys()))
            matched_death_pids = death_lookup_result.get("matched", set())
            central_death_lookup_available = bool(death_lookup_result.get("available"))
            central_death_lookup_message = death_lookup_result.get("message", "")
            for pid, candidate_rows in central_death_candidates.items():
                found = pid in matched_death_pids
                for candidate_row in candidate_rows:
                    if central_death_lookup_available:
                        candidate_row["เทียบตาย"] = "พบข้อมูล" if found else "ไม่พบข้อมูล"
                    else:
                        candidate_row["เทียบตาย"] = "ใช้ไม่ได้"
                    candidate_row["_central_death_mismatch"] = found
                    if found:
                        central_death_mismatch_count += 1

        return {
            "data": transformed_data,
            "columns": person_columns + columns,
            "matched_count": matched_count,
            "unmatched_count": unmatched_count,
            "pid_matched_count": pid_matched_count,
            "pid_unmatched_count": len(transformed_data) - pid_matched_count,
            "cid_matched_count": cid_matched_count,
            "cid_unmatched_count": unmatched_count,
            "has_discharge": bool(discharge_column),
            "central_death_mismatch_count": central_death_mismatch_count,
            "central_death_lookup_available": central_death_lookup_available,
            "central_death_lookup_message": central_death_lookup_message
        }

    except Exception as e:
        raise Exception(f"เกิดข้อผิดพลาดในการแปลงข้อมูล: {e}")


def export_excel(data: list, columns: list, original_filename: str) -> str:
    """
    ส่งออกข้อมูลที่แปลงแล้วเป็นไฟล์ Excel ใหม่

    Args:
        data: list ของ dict ข้อมูลที่แปลงแล้ว
        columns: list ของชื่อคอลัมน์
        original_filename: ชื่อไฟล์ต้นฉบับ

    Returns:
        str: path ไปยังไฟล์ Excel ที่สร้าง
    """
    try:
        # สร้างโฟลเดอร์ uploads หากยังไม่มี
        os.makedirs(UPLOADS_DIR, exist_ok=True)

        # สร้าง workbook ใหม่
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # กำหนด style สำหรับ header
        header_font = Font(name="TH SarabunPSK", size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        export_columns = [col for col in columns if not str(col).startswith("_")]

        # เขียน header
        for col_idx, col_name in enumerate(export_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # กำหนด style สำหรับ data
        data_font = Font(name="TH SarabunPSK", size=14)
        data_alignment = Alignment(vertical="center")

        # เขียนข้อมูล
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(export_columns, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # Auto-size columns
        for col_idx, col_name in enumerate(export_columns, 1):
            max_length = len(str(col_name))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass
            # กำหนดความกว้างคอลัมน์ (เพิ่ม padding)
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

        # ตั้งแถวแรกเป็น freeze pane
        ws.freeze_panes = "A2"

        # สร้างชื่อไฟล์ output
        name_without_ext = os.path.splitext(original_filename)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{name_without_ext}_completed_{timestamp}.xlsx"
        output_path = os.path.join(UPLOADS_DIR, output_filename)

        # บันทึกไฟล์
        wb.save(output_path)
        wb.close()

        return output_path

    except Exception as e:
        raise Exception(f"เกิดข้อผิดพลาดในการส่งออกไฟล์ Excel: {e}")
